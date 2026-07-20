from __future__ import annotations

import re
from datetime import date, time, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..._time import pacific_today
from ...models import DAY_ORDER
from ..errors import DirectSourceError
from ..parsing import _parse_clock_time, _parse_hours_range, _payload, _resolve_yearless_date, _session


_WEEKDAYS = {day.title() for day in DAY_ORDER[:5]}
_KNOWN_SHEETS = _WEEKDAYS | {"Weekend", "Long Course Notice"}
_CLOSED_BANNER = re.compile(r"^closed\b[\s:,-]*(.*)", re.IGNORECASE)


def _extract_koret(path: Path) -> dict:
    workbook = load_workbook(path, data_only=True)
    visible_sheets = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
    unknown = sorted(sheet.title for sheet in visible_sheets if sheet.title not in _KNOWN_SHEETS)
    if unknown:
        raise DirectSourceError(f"Unclassified visible Koret sheet(s): {', '.join(unknown)}")

    sessions: list[dict] = []
    closures: list[dict] = []
    for sheet in visible_sheets:
        if sheet.title in _WEEKDAYS:
            sessions.extend(_weekday_sessions(sheet))
            banner = _closed_banner_closure(sheet)
            if banner is not None:
                closures.append(banner)
        elif sheet.title == "Weekend":
            weekend_sessions, weekend_closures = _weekend_schedule(sheet)
            sessions.extend(weekend_sessions)
            closures.extend(weekend_closures)
        elif sheet.title == "Long Course Notice":
            closures.extend(_notice_closures(sheet))

    return _payload("pool_hours", sessions, closures=closures)


def _weekday_sessions(sheet: Worksheet) -> list[dict]:
    hours = _headline_hours(sheet, start_row=1, end_row=sheet.max_row)
    if hours is None:
        hours = _time_grid_range(sheet, start_row=1, end_row=sheet.max_row)
    if hours is None:
        if _sheet_text(sheet).strip():
            return []
        raise DirectSourceError(f"{sheet.title}: no hours or time grid found")
    start, end, evidence = hours
    windows = _subtract_windows(start, end, _merged_closed_windows(sheet, fallback_end=end))
    return [
        _session(sheet.title.lower(), "lap_swim", window_start, window_end, evidence)
        for window_start, window_end in windows
    ]


def _weekend_schedule(sheet: Worksheet) -> tuple[list[dict], list[dict]]:
    sunday_row = next(
        (cell.row for row in sheet.iter_rows() for cell in row if str(cell.value or "").strip().lower() == "sunday"),
        None,
    )
    if sunday_row is None:
        raise DirectSourceError("Weekend: Sunday marker not found")

    sessions: list[dict] = []
    saturday = _headline_hours(sheet, start_row=1, end_row=sunday_row - 1, label="Saturday")
    if saturday is None:
        saturday = _time_grid_range(sheet, start_row=1, end_row=sunday_row - 1)
    if saturday is not None:
        start, end, evidence = saturday
        sessions.append(_session("saturday", "lap_swim", start, end, evidence))

    sunday_text = " ".join(
        str(sheet.cell(row, column).value or "")
        for row in range(sunday_row, sheet.max_row + 1)
        for column in range(1, sheet.max_column + 1)
    )
    if "closed" not in sunday_text.lower():
        sunday = _headline_hours(sheet, start_row=sunday_row, end_row=sheet.max_row, label="Sunday")
        if sunday is None:
            sunday = _time_grid_range(sheet, start_row=sunday_row + 1, end_row=sheet.max_row)
        if sunday is not None:
            start, end, evidence = sunday
            sessions.append(_session("sunday", "lap_swim", start, end, evidence))

    closures = _ordinal_closures(sheet)
    return sessions, closures


def _headline_hours(
    sheet: Worksheet, *, start_row: int, end_row: int, label: str | None = None
) -> tuple[str, str, str] | None:
    """The stated opening hours, which are authoritative over the lane grid: the
    grid runs past closing on days the USF teams have the water. The colon is
    optional because the Weekend sheet writes "Hours 8am-4pm" without one, and
    "(Summer Hours)" appears inside banners carrying no range at all — so a cell
    only counts once a range actually parses out of it."""
    for row in range(start_row, end_row + 1):
        for column in range(1, sheet.max_column + 1):
            value = str(sheet.cell(row, column).value or "").strip()
            if "hours" not in value.lower():
                continue
            try:
                start, end = _parse_hours_range(value)
            except DirectSourceError:
                continue
            return start, end, f"{label or sheet.title} {value}"
    return None


def _time_grid_range(sheet: Worksheet, *, start_row: int, end_row: int) -> tuple[str, str, str] | None:
    times = [
        _clock(sheet.cell(row, 1).value)
        for row in range(start_row, end_row + 1)
        if _clock(sheet.cell(row, 1).value) is not None
    ]
    if not times:
        return None
    return times[0], _shift_hour(times[-1]), f"{sheet.title} time grid"


def _merged_closed_windows(sheet: Worksheet, *, fallback_end: str | None = None) -> list[tuple[str, str]]:
    windows: list[tuple[str, str]] = []
    for row in sheet.iter_rows():
        for cell in row:
            if "closed" not in str(cell.value or "").lower():
                continue
            merged = next((area for area in sheet.merged_cells.ranges if cell.coordinate in area), None)
            if merged is None or merged.min_col > 2 or merged.max_col < 9:
                continue
            start = _clock(sheet.cell(merged.min_row, 1).value)
            # A closed block reaching the last grid row has no next-row time
            # cell to read the end from; clamp to the sheet's closing hour.
            end = _clock(sheet.cell(merged.max_row + 1, 1).value) or fallback_end
            if start is not None and end is not None and start < end:
                windows.append((start, end))
    return windows


def _closed_banner_closure(sheet: Worksheet) -> dict | None:
    """A full-width merged banner starting with "Closed" (e.g. "Closed
    Juneteenth" over A2:S2) marks the whole weekday closed even though the
    lane grid below it stays populated."""
    for merged in sheet.merged_cells.ranges:
        if merged.min_col != 1 or merged.max_col < 9:
            continue
        match = _CLOSED_BANNER.match(str(sheet.cell(merged.min_row, 1).value or "").strip())
        if match is None:
            continue
        today = pacific_today()
        weekday = DAY_ORDER.index(sheet.title.lower())
        closed_date = today + timedelta(days=(weekday - today.weekday()) % 7)
        return {
            "start": closed_date.isoformat(),
            "end": closed_date.isoformat(),
            "reason": match.group(1).strip() or "Koret closed",
        }
    return None


def _notice_closures(sheet: Worksheet) -> list[dict]:
    text = _sheet_text(sheet)
    closures: list[dict] = []
    for month, day, start, end in re.findall(
        r"on\s+(\d{1,2})/(\d{1,2}).*?closed\s+from\s+(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    ):
        closure_date = _resolve_yearless_date(int(month), int(day))
        closures.append({
            "start": closure_date.isoformat(),
            "end": closure_date.isoformat(),
            "start_time": _normalize_clock(start),
            "end_time": _normalize_clock(end),
            "reason": "Change from long course to short course",
        })
    return closures


def _ordinal_closures(sheet: Worksheet) -> list[dict]:
    text = _sheet_text(sheet)
    closures: list[dict] = []
    today = pacific_today()
    for day_text in re.findall(r"Saturday\s+CLOSED\s+(\d{1,2})(?:ST|ND|RD|TH)", text, flags=re.IGNORECASE):
        day_number = int(day_text)
        # The note names a Saturday by day-of-month only. Resolve it to this
        # month or next, but only accept a date that actually falls on a
        # Saturday and isn't stale — otherwise the note refers to a week
        # the sheet no longer covers and gets dropped.
        for month_offset in (0, 1):
            month = (today.month - 1 + month_offset) % 12 + 1
            year = today.year + (today.month - 1 + month_offset) // 12
            try:
                candidate = date(year, month, day_number)
            except ValueError:
                continue
            if candidate.weekday() == 5 and candidate >= today - timedelta(days=7):
                closures.append({"start": candidate.isoformat(), "end": candidate.isoformat(), "reason": "Koret closed"})
                break
    return closures


def _sheet_text(sheet: Worksheet) -> str:
    return "\n".join(str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value not in (None, ""))


def _clock(value) -> str | None:
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if value in (None, ""):
        return None
    try:
        return _parse_clock_time(str(value))
    except DirectSourceError:
        return None


def _normalize_clock(value: str) -> str:
    hour, minute = (int(part) for part in value.split(":"))
    return f"{hour:02d}:{minute:02d}"


def _shift_hour(value: str) -> str:
    hour, minute = (int(part) for part in value.split(":"))
    return f"{(hour + 1) % 24:02d}:{minute:02d}"


def _subtract_windows(start: str, end: str, excluded: list[tuple[str, str]]) -> list[tuple[str, str]]:
    windows: list[tuple[str, str]] = []
    cursor = start
    for excluded_start, excluded_end in sorted(excluded):
        if excluded_end <= cursor or excluded_start >= end:
            continue
        if cursor < excluded_start:
            windows.append((cursor, min(excluded_start, end)))
        cursor = max(cursor, excluded_end)
    if cursor < end:
        windows.append((cursor, end))
    return windows
