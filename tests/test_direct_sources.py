from __future__ import annotations

import hashlib
import io
import zipfile
from datetime import date, time, timedelta

import pytest
from openpyxl import Workbook

import schedules.direct_sources as direct_sources
from schedules.direct_sources import (
    DirectSourceError,
    _cache_text,
    _xlsx_content_sha256,
    _extract_24_hour_fitness,
    _extract_city_sports,
    _extract_equinox,
    _extract_fitness_sf,
    _extract_jccsf,
    _extract_koret,
    _extract_pomeroy,
    _extract_sfsu_aquatics,
    _extract_ucsf_bakar,
    _extract_ucsf_fitness,
    _extract_ymca_location,
)


def test_cache_text_matches_original_response_bytes_with_crlf(tmp_path):
    text = "first\r\nsecond\r\n"
    sha256 = hashlib.sha256(text.encode("utf-8")).hexdigest()

    _cache_text(tmp_path, sha256, "html", text)
    path, from_cache = _cache_text(tmp_path, sha256, "html", text)

    assert path.read_bytes() == text.encode("utf-8")
    assert from_cache is True


def test_cache_text_can_key_dynamic_html_by_semantic_fingerprint(tmp_path):
    sha256 = hashlib.sha256(b"same extracted hours").hexdigest()

    original, _ = _cache_text(tmp_path, sha256, "html", "<p>hours</p><script>nonce-a</script>")
    cached, from_cache = _cache_text(tmp_path, sha256, "html", "<p>hours</p><script>nonce-b</script>")

    assert cached == original
    assert from_cache is True
    assert (cached.parent / "source.sha256").read_text().strip() == sha256


def test_xlsx_identity_ignores_zip_metadata():
    def make_xlsx(timestamp):
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w") as archive:
            info = zipfile.ZipInfo("xl/workbook.xml", date_time=timestamp)
            archive.writestr(info, b"<workbook/>")
        return output.getvalue()

    assert _xlsx_content_sha256(make_xlsx((2026, 7, 10, 10, 0, 0))) == _xlsx_content_sha256(
        make_xlsx((2026, 7, 10, 11, 0, 0))
    )


def test_jccsf_html_extractor_models_lap_and_family_hours():
    payload = _extract_jccsf(
        """
        <h3>Aquatics Center Hours</h3>
        <p>Monday – Friday: 5:30 am – 9:45 pm</p>
        <p>Saturday & Sunday: 7:00 am – 6:45 pm*</p>
        <h3>Rec Pool Hours</h3>
        <p>Monday & Wednesday: 5:30 am – Noon, 3:00 – 9:45 pm</p>
        <p>Tuesday: 5:30 – 11:30 am, 3:00 – 9:45 pm</p>
        <p>Thursday: 5:30 am – Noon, 3:00 – 9:45 pm</p>
        <p>Friday: 5:30 – 10:00 am, 1:30 – 9:45 pm</p>
        <p>Saturday & Sunday: 7:00 – 8:00 am, 2:00 – 6:45 pm</p>
        <p>The Lap Pool is available for lap swimming during Aquatics Center hours.</p>
        """
    )

    assert any(
        s["day"] == "monday" and s["type"] == "lap_swim" and s["start"] == "05:30" and s["end"] == "21:45"
        for s in payload["sessions"]
    )
    assert any(s["type"] == "family_swim" for s in payload["sessions"])
    assert any(
        s["day"] == "tuesday"
        and s["type"] == "family_swim"
        and s["start"] == "15:00"
        and s["end"] == "21:45"
        for s in payload["sessions"]
    )


def test_jccsf_html_extractor_rejects_page_when_posted_hours_change():
    with pytest.raises(DirectSourceError, match="Expected source text not found"):
        _extract_jccsf(
            """
            <h3>Aquatics Center Hours</h3>
            <p>Monday – Friday: 6:00 am – 9:00 pm</p>
            <p>The Lap Pool is available for lap swimming during Aquatics Center hours.</p>
            """
        )


def _koret_workbook(tmp_path, sheets):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in sheets.items():
        sheet = workbook.create_sheet(title)
        for row in rows:
            sheet.append(row)
    path = tmp_path / "koret.xlsx"
    workbook.save(path)
    return path


def test_koret_google_sheet_extractor_reads_weekday_and_weekend_hours(tmp_path):
    sheets = {
        day: [[day], ["Hours: 6am - 9pm"], [time(6), "slow"], [time(20), "slow"]]
        for day in ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday")
    }
    sheets["Weekend"] = [[None, "Saturday"], [time(8), "slow"], [time(17), "slow"], ["Sunday"], [time(8), "slow"], [time(17), "slow"]]
    sheets["Long Course Notice"] = [["Long Course Notice"]]
    payload = _extract_koret(_koret_workbook(tmp_path, sheets))

    assert any(s["day"] == "monday" and s["start"] == "06:00" and s["end"] == "21:00" for s in payload["sessions"])
    assert any(s["day"] == "saturday" and s["start"] == "08:00" and s["end"] == "18:00" for s in payload["sessions"])
    assert any(s["day"] == "sunday" and s["start"] == "08:00" and s["end"] == "18:00" for s in payload["sessions"])


def test_koret_weekend_stated_hours_win_over_the_lane_grid(tmp_path):
    """The Weekend sheet states "Hours 8am-4pm" with no colon, over a grid whose
    lanes stay assigned until 5pm for the USF squads. Closing time is 4pm; the
    Sunday banner carries "(Summer Hours)" but no range and must not be read as
    one."""
    sheets = {
        day: [[day], ["Hours: 7am-7pm (Summer Hours)"], [time(6), "slow"], [time(20), "slow"]]
        for day in ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday")
    }
    sheets["Weekend"] = [
        [None, "Saturday"],
        [None, "Hours 8am-4pm (Summer Hours)"],
        [time(8), "slow"],
        [time(17), "slow"],
        ["Sunday"],
        ["KORET CLOSED (Summer Hours)"],
        [time(8), "slow"],
        [time(17), "slow"],
    ]
    payload = _extract_koret(_koret_workbook(tmp_path, sheets))

    saturday = [s for s in payload["sessions"] if s["day"] == "saturday"]
    assert [(s["start"], s["end"]) for s in saturday] == [("08:00", "16:00")]
    assert "8am-4pm" in saturday[0]["evidence"]
    assert not [s for s in payload["sessions"] if s["day"] == "sunday"]


def test_koret_google_sheet_extractor_splits_hours_around_closed_grid_rows(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Monday"
    sheet.append(["Monday"])
    sheet.append(["Hours: 7am-7pm"])
    sheet.append([])
    sheet.append([None, "Lane 1", "Lane 2", "Lane 3", "Lane 4", "Lane 5", "Lane 6", "Lane 7", "Lane 8"])
    sheet.append([time(7), "fast"])
    sheet.append([time(8), "fast"])
    sheet.append([])
    sheet.append([time(9), "Closed for bulk head transition"])
    sheet.append([time(10)])
    sheet.append([time(11), "fast"])
    sheet.merge_cells("B8:I9")
    for day in ("Tuesday", "Wednesday", "Thursday", "Friday"):
        other = workbook.create_sheet(day)
        other.append([day])
        other.append(["Hours: 7am-7pm"])
    weekend = workbook.create_sheet("Weekend")
    weekend.append([None, "Saturday"])
    weekend.append(["Sunday"])
    workbook.create_sheet("Long Course Notice")
    path = tmp_path / "koret.xlsx"
    workbook.save(path)
    payload = _extract_koret(path)

    assert [(session["start"], session["end"]) for session in payload["sessions"] if session["day"] == "monday"] == [
        ("07:00", "09:00"),
        ("11:00", "19:00"),
    ]


def test_koret_google_sheet_extractor_reads_dated_notice_closure(tmp_path):
    sheets = {
        day: [[day], ["Hours: 7am-7pm"]]
        for day in ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday")
    }
    sheets["Weekend"] = [[None, "Saturday"], ["Sunday"]]
    sheets["Long Course Notice"] = [["Long Course Notice"], ["On 7/13 the pool will be closed from 8:30-10:30 to change back to short course."]]
    payload = _extract_koret(_koret_workbook(tmp_path, sheets))

    assert payload["closures"] == [{
        "start": "2026-07-13",
        "end": "2026-07-13",
        "start_time": "08:30",
        "end_time": "10:30",
        "reason": "Change from long course to short course",
    }]


def test_koret_notice_closure_rolls_year_forward(tmp_path, monkeypatch):
    monkeypatch.setattr(direct_sources, "pacific_today", lambda: date(2026, 12, 20))
    sheets = {
        day: [[day], ["Hours: 7am-7pm"]]
        for day in ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday")
    }
    sheets["Weekend"] = [[None, "Saturday"], ["Sunday"]]
    sheets["Long Course Notice"] = [["Long Course Notice"], ["On 1/5 the pool will be closed from 8:30-10:30 to change back to short course."]]
    payload = _extract_koret(_koret_workbook(tmp_path, sheets))

    assert payload["closures"][0]["start"] == "2027-01-05"


def test_koret_closed_banner_weekday_emits_closure_and_keeps_sessions(tmp_path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for day in ("Monday", "Tuesday", "Wednesday", "Thursday"):
        sheet = workbook.create_sheet(day)
        sheet.append([day])
        sheet.append(["Hours: 7am-7pm"])
    friday = workbook.create_sheet("Friday")
    friday.append(["Friday"])
    friday.append(["Closed Juneteenth"])
    friday.append(["Hours: 7am-7pm"])
    friday.merge_cells("A2:I2")
    weekend = workbook.create_sheet("Weekend")
    weekend.append([None, "Saturday"])
    weekend.append(["Sunday"])
    workbook.create_sheet("Long Course Notice")
    path = tmp_path / "koret.xlsx"
    workbook.save(path)

    payload = _extract_koret(path)

    today = direct_sources.pacific_today()
    expected = (today + timedelta(days=(4 - today.weekday()) % 7)).isoformat()
    assert {"start": expected, "end": expected, "reason": "Juneteenth"} in payload["closures"]
    assert any(session["day"] == "friday" for session in payload["sessions"])


def test_koret_ordinal_closures_only_accept_real_saturdays(tmp_path):
    sheets = {
        day: [[day], ["Hours: 7am-7pm"]]
        for day in ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday")
    }
    sheets["Weekend"] = [[None, "Saturday"], [time(8), "slow"], [time(16), "slow"], ["Sunday"], ["Saturday CLOSED 31ST"]]
    payload = _extract_koret(_koret_workbook(tmp_path, sheets))

    for closure in payload["closures"]:
        assert date.fromisoformat(closure["start"]).weekday() == 5


def test_koret_closed_block_reaching_last_grid_row_is_subtracted(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Monday"
    sheet.append(["Monday"])
    sheet.append(["Hours: 7am-7pm"])
    sheet.append([None, "Lane 1", "Lane 2", "Lane 3", "Lane 4", "Lane 5", "Lane 6", "Lane 7", "Lane 8"])
    sheet.append([time(7), "fast"])
    sheet.append([time(8), "fast"])
    sheet.append([time(9), "Closed for cleaning"])
    sheet.append([time(10)])
    sheet.merge_cells("B6:I7")
    for day in ("Tuesday", "Wednesday", "Thursday", "Friday"):
        other = workbook.create_sheet(day)
        other.append([day])
        other.append(["Hours: 7am-7pm"])
    weekend = workbook.create_sheet("Weekend")
    weekend.append([None, "Saturday"])
    weekend.append(["Sunday"])
    workbook.create_sheet("Long Course Notice")
    path = tmp_path / "koret.xlsx"
    workbook.save(path)

    payload = _extract_koret(path)

    assert [(session["start"], session["end"]) for session in payload["sessions"] if session["day"] == "monday"] == [
        ("07:00", "09:00"),
    ]


def test_pomeroy_html_extractor_handles_table_rowspans():
    payload = _extract_pomeroy(
        """
        <h1>Upcoming Pool Closure Dates:</h1>
        <p><span>Monday, May 25th - Memorial Day</span></p>
        <p><span>Friday, June 19th - Juneteenth</span></p>
        <table class="PoolSchedule">
          <thead><tr><th>Monday</th><th>Tuesday</th><th>Wednesday</th></tr></thead>
          <tbody>
            <tr>
              <td rowspan="2">1pm - 2:55pm<br><span>Open Swim</span></td>
              <td>8am - 8:55am<br><span>Lap Swim</span></td>
              <td>8am - 8:55am<br><span>Lap Swim</span></td>
            </tr>
            <tr>
              <td>9am - 9:55am<br><span>Open Swim</span></td>
              <td>6pm - 6:55pm<br><span>Lap Swim</span></td>
            </tr>
          </tbody>
        </table>
        """
    )

    assert {
        (s["day"], s["type"], s["start"], s["end"])
        for s in payload["sessions"]
    } == {
        ("monday", "family_swim", "13:00", "14:55"),
        ("tuesday", "lap_swim", "08:00", "08:55"),
        ("tuesday", "family_swim", "09:00", "09:55"),
        ("wednesday", "lap_swim", "08:00", "08:55"),
        ("wednesday", "lap_swim", "18:00", "18:55"),
    }
    assert [closure["reason"] for closure in payload["closures"]] == ["Memorial Day", "Juneteenth"]


def test_24_hour_fitness_extractor_reads_access_hours():
    payload = _extract_24_hour_fitness(
        """
        <h2>Gym Hours</h2>
        <span class="ih-days">Monday - Thursday</span>
        <span class="ih-hours">05:00 AM - 10:00 PM</span>
        <span class="ih-days">Friday</span>
        <span class="ih-hours">05:00 AM - 09:00 PM</span>
        <span class="ih-days">Saturday - Sunday</span>
        <span class="ih-hours">06:00 AM - 08:00 PM</span>
        """
    )

    assert payload["sessions"] == []
    assert payload["schedule_basis"] == "facility_hours"
    assert len(payload["access_hours"]) == 7
    assert any(a["day"] == "thursday" and a["end"] == "22:00" for a in payload["access_hours"])


def test_24_hour_fitness_extractor_models_temporary_closure():
    payload = _extract_24_hour_fitness(
        """
        <h2>Gym Hours</h2>
        <p>Under Renovation This club is temporarily closed for renovation.
        We can't wait to welcome you back on 05/23/2026 when our improvements are complete.</p>
        """
    )

    assert payload["schedule_basis"] == "temporarily_closed"
    assert payload["sessions"] == []
    assert payload["access_hours"] == []
    assert payload["closures"][0]["reason"] == "Temporarily closed for renovation"


def test_24_hour_fitness_extractor_models_reopened_facility_hours():
    payload = _extract_24_hour_fitness(
        """
        <h2>Recently Renovated!</h2>
        <p>Reimagined. Reopened. Ready for you.</p>
        <p>Indoor Lap Pool</p>
        <h2>Gym Hours</h2>
        <span class="ih-days">Monday</span>
        <span class="ih-hours">05:00 AM - 11:59 PM</span>
        <span class="ih-days">Tuesday - Thursday</span>
        <span class="ih-hours">12:00 AM - 11:59 PM</span>
        <span class="ih-days">Friday</span>
        <span class="ih-hours">12:00 AM - 09:00 PM</span>
        <span class="ih-days">Saturday - Sunday</span>
        <span class="ih-hours">05:00 AM - 09:00 PM</span>
        """
    )

    assert payload["schedule_basis"] == "facility_hours"
    assert payload["sessions"] == []
    assert payload["closures"] == []
    assert len(payload["access_hours"]) == 7
    assert any(a["day"] == "monday" and a["end"] == "23:59" for a in payload["access_hours"])
    assert any(a["day"] == "wednesday" and a["start"] == "00:00" for a in payload["access_hours"])


def test_city_sports_extractor_reads_club_hours():
    payload = _extract_city_sports(
        """
        <h1>SAN FRANCISCO - 20TH AVE</h1>
        <p>lap pool</p>
        <p>HOURS Mon - Thu 5:00am - 11:00pm Fri 5:00am - 10:00pm
        Sat - Sun 8:00am - 8:00pm Special Club Hours</p>
        """
    )

    assert payload["schedule_basis"] == "facility_hours"
    assert len(payload["access_hours"]) == 7
    assert any(a["day"] == "friday" and a["end"] == "22:00" for a in payload["access_hours"])


def test_equinox_extractor_reads_schema_hours():
    payload = _extract_equinox(
        """
        <h1>Equinox Sports Club San Francisco</h1>
        <p>Indoor Pool</p>
        "openingHoursSpecification": [
          {"dayOfWeek": ["Monday","Tuesday"],"opens": "05:00","closes": "22:00"},
          {"dayOfWeek": ["Saturday","Sunday"],"opens": "07:00","closes": "18:00"}
        ]
        """
    )

    assert payload["schedule_basis"] == "facility_hours"
    assert any(a["day"] == "monday" and a["start"] == "05:00" for a in payload["access_hours"])
    assert any(a["day"] == "sunday" and a["end"] == "18:00" for a in payload["access_hours"])


def test_fitness_sf_extractor_reads_pool_hours_from_location_hours():
    payload = _extract_fitness_sf(
        """
        <h1>FITNESS SF Fillmore</h1>
        <p>25-yard, 5-lane swimming pool</p>
        <p>Mon - Thu: 5 am - 12 am Fri: 5 am - 11 pm Sat - Sun: 7 am - 8 pm 1-415-348-6377</p>
        """
    )

    assert payload["schedule_basis"] == "pool_hours"
    assert any(a["day"] == "thursday" and a["end"] == "23:59" for a in payload["access_hours"])


def test_sfsu_extractor_reads_natatorium_hours():
    payload = _extract_sfsu_aquatics(
        """
        <h2>Natatorium Hours of Operation</h2>
        <p>Mon, Wed, Thur: Noon - 4:00 p.m. Tue, Fri: 10:00am- 1:30pm Saturday/ Sunday: Closed</p>
        <p>Lap Pool: six lanes.</p>
        """
    )

    assert payload["schedule_basis"] == "pool_hours"
    assert len(payload["access_hours"]) == 5
    assert any(a["day"] == "monday" and a["start"] == "12:00" for a in payload["access_hours"])
    assert any(a["day"] == "friday" and a["start"] == "10:00" for a in payload["access_hours"])


def test_ucsf_bakar_extractor_reads_facility_hours():
    payload = _extract_ucsf_bakar(
        """
        <p>Facility Hours: Monday-Friday, 6:00 am-9:00 pm;
        Saturday-Sunday, 8:00 am-6:00 pm</p>
        """
    )

    assert payload["schedule_basis"] == "facility_hours"
    assert len(payload["access_hours"]) == 7
    assert any(a["day"] == "monday" and a["start"] == "06:00" for a in payload["access_hours"])
    assert any(a["day"] == "sunday" and a["end"] == "18:00" for a in payload["access_hours"])


def test_ucsf_fitness_extractor_handles_millberry_page():
    payload = _extract_ucsf_fitness(
        """
        <p>Facility Hours: Monday-Friday, 6:00 am-9:00 pm;
        Saturday-Sunday, 8:00 am-4:00 pm Millberry Union</p>
        """
    )

    assert payload["schedule_basis"] == "facility_hours"
    assert any(a["day"] == "sunday" and a["end"] == "16:00" for a in payload["access_hours"])


def test_ymca_extractor_reads_first_location_hours_block():
    payload = _extract_ymca_location(
        """
        <h2>Letterman Pool &amp; Gym Hours</h2>
        <div class="tr-accordion_day-hour__list"><span>Monday</span><span>5:30 am – 8:30 pm</span></div>
        <div class="tr-accordion_day-hour__list"><span>Tuesday</span><span>5:30 am – 8:30 pm</span></div>
        <div class="tr-accordion_day-hour__list"><span>Wednesday</span><span>5:30 am – 8:30 pm</span></div>
        <div class="tr-accordion_day-hour__list"><span>Thursday</span><span>5:30 am – 8:30 pm</span></div>
        <div class="tr-accordion_day-hour__list"><span>Friday</span><span>5:30 am – 8:30 pm</span></div>
        <div class="tr-accordion_day-hour__list"><span>Saturday</span><span>7:00 am – 4:30 pm</span></div>
        <div class="tr-accordion_day-hour__list"><span>Sunday</span><span>Closed</span></div>
        <h2>Other Hours</h2>
        <div class="tr-accordion_day-hour__list"><span>Monday</span><span>1:00 pm – 2:00 pm</span></div>
        """
    )

    assert payload["schedule_basis"] == "facility_hours"
    assert len(payload["access_hours"]) == 6
    assert any(a["day"] == "saturday" and a["end"] == "16:30" for a in payload["access_hours"])
    assert not any(a["day"] == "sunday" for a in payload["access_hours"])


def test_ymca_extractor_prefers_facility_hours_block_with_day_ranges(monkeypatch):
    monkeypatch.setattr(direct_sources, "pacific_today", lambda: date(2026, 5, 17))

    payload = _extract_ymca_location(
        """
        <h2>Facility Hours</h2>
        <div class="tr-accordion_day-hour__list"><span>Monday-Friday</span><span>6:30 a.m. – 7:45 p.m.</span></div>
        <div class="tr-accordion_day-hour__list"><span>Saturday and Sunday</span><span>8:00 a.m. – 3:45 p.m.</span></div>
        <h2>Holiday Hours</h2>
        <p>Monday, May 25 (Memorial Day)</p>
        <p>8:00 a.m. – 1:30 p.m.</p>
        <h4>Contact</h4>
        <div class="tr-accordion_day-hour__list"><span>Monday</span><span>8:00 am – 6:00 pm</span></div>
        """
    )

    assert len(payload["access_hours"]) == 7
    assert any(a["day"] == "monday" and a["start"] == "06:30" and a["end"] == "19:45" for a in payload["access_hours"])
    assert any(a["day"] == "sunday" and a["start"] == "08:00" and a["end"] == "15:45" for a in payload["access_hours"])
    assert payload["access_exceptions"] == [{
        "date": "2026-05-25",
        "evidence": "Monday, May 25 (Memorial Day) 8:00 a.m. – 1:30 p.m.",
        "label": "Holiday facility hours",
        "reason": "Memorial Day",
        "start": "08:00",
        "end": "13:30",
    }]


def test_ymca_extractor_uses_pool_hours_when_page_gives_pool_rule(monkeypatch):
    monkeypatch.setattr(direct_sources, "pacific_today", lambda: date(2026, 5, 17))

    payload = _extract_ymca_location(
        """
        <h2>Hours</h2>
        <div class="tr-accordion_day-hour__list"><span>Monday</span><span>5:30 am – 9:00 pm</span></div>
        <div class="tr-accordion_day-hour__list"><span>Tuesday</span><span>5:30 am – 9:00 pm</span></div>
        <h2>Holiday Hours</h2>
        <p>Monday, May 25 (Memorial Day)</p>
        <p>7:00 a.m. – 2:00 p.m.</p>
        <p>Pool Closes at 1:30 p.m.</p>
        <p>Pool Hours Opens 30 min after, closes 30 min before facility</p>
        """
    )

    assert payload["schedule_basis"] == "pool_hours"
    assert payload["access_hours"][0]["start"] == "06:00"
    assert payload["access_hours"][0]["end"] == "20:30"
    assert payload["access_exceptions"] == [{
        "date": "2026-05-25",
        "evidence": "Monday, May 25 (Memorial Day) 7:00 a.m. – 2:00 p.m. Pool Closes at 1:30 p.m.",
        "label": "Holiday pool hours",
        "reason": "Memorial Day",
        "start": "07:30",
        "end": "13:30",
    }]
